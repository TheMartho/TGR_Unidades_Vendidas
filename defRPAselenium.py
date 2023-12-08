from RPA.Browser.Selenium import Selenium;
from RPA.Excel.Application import Application
from RPA.Windows import Windows
from RPA.Excel.Files import Files;
import time
import random
from datetime import timedelta
import os
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By



listSCRAPIADO= ([{
               'CUOTA':"",
               'NRO FOLIO':"", 
               'VALOR':"",
               'VENCIMIENTO':"",
                'TOTA A PAGAR':"",
                 }])

listSFormato= ([{
               'pathubicacion':"",
               'Nombre Solicitante':"", 
               'fecha':"",
               'gerente':"",
                'Rut':"",
                'Monto':"",
                'RUTtesoria':"",
                'Direccio':"",
                'Glosagasto':"",
                'Detallegasto':"",
                'CentroGestion':"",
                'Contribuciones':"",
                 }])

browser = Selenium()
library = Windows() 
lib = Files()
app = Application()
TxExcel=int(680)
RutExcel="-"
InmobiliariaExcel="-"
Unidad="-"

def Pyasset(asset):
    lib.open_workbook("PyAsset\Config.xlsx")      #ubicacion del libro
    lib.read_worksheet("Variables")       #nombre de la hoja
    config=lib.read_worksheet_as_table(name='Variables',header=True, start=1).data
    for x in config:
        if x[0]==asset:
            exitdato= str(x[1])
        
            return exitdato

def openweb(u):
    duracion=timedelta(seconds=59)
    browser.set_selenium_page_load_timeout(duracion)
    try:
        browser.open_available_browser(u)
        browser.maximize_browser_window() 
    except:
         browser.reload_page()


    validacion= browser.get_text("//DIV[@class='dentro_letra'][text()='Contribuciones']")
    if validacion == 'Contribuciones': print("ingresando a "+validacion) 
    state_tgc_Inicio=True
 
    time.sleep(random.uniform(5,7))

def clickweb(elemento):
    time.sleep(random.uniform(1,2))
    browser.click_element(elemento)
    time.sleep(random.uniform(1,2))

def typeinputText(elemento,texto):
    time.sleep(random.uniform(1,2))
    browser.input_text(elemento,texto)
    time.sleep(random.uniform(1,2))

def obtenertabla(elemento,columna,celdas):
    time.sleep(random.uniform(1,2))
    browser.get_table_cell(locator=elemento,column=columna,row=celdas)
    time.sleep(random.uniform(1,2))

def obtenerTexto(elemento):
    time.sleep(random.uniform(1,2))
    browser.get_text(elemento)
    time.sleep(random.uniform(1,2))

def tiempoespera():
    time.sleep(random.uniform(21,25))

def cerraNavegador():
    #browser.close_browser()
    browser.close_all_browsers()
    print("----------------------proceso terminado----------------------")

def destacar(elemento):
    browser.highlight_elements(elemento)
    time.sleep(random.uniform(3,7))

def LOGconsulta(Región,Comuna,RolMatriz,Rol):
    print('----------------------Consultado-----------------------------')
    print('region = '+str(Región))
    print('Comuna = '+str(Comuna))
    print('Rol Matriz = '+str(RolMatriz))
    print('Rol = '+str(Rol))

def extraertablita():

    
    print(browser.get_text("//DIV[@id='example_info']/self::DIV"))
    
    scraping=browser.get_text("//TABLE[@id='example']")
    #recorrerFilasDescargas()
    print(scraping)
    return scraping

def filtrarCuota():
    #Noviembre trae 4 / Septiembre trae 3 / Junio trae 2 / Abril trae 1
    #Metodo para saber en que mes se esta corriendo y descargar los PDF´s que coincidan
    fecha_actual=datetime.now()
    fecha_formateada = fecha_actual.strftime("%B")
    anio_actual = fecha_actual.strftime('%Y')

    if fecha_formateada=="November":
        return "4-"+anio_actual
    elif fecha_formateada=="September":
        return "3-"+anio_actual
    elif fecha_formateada=="June":
        return "2-"+anio_actual
    elif fecha_formateada=="April":
        return "1-"+anio_actual
    elif fecha_formateada=="December":
        return "9-"+anio_actual
    else:
        print("Opción no valida")



def valoresReporte(rut,inmobiliaria,unidad):
     global RutExcel
     global InmobiliariaExcel
     global TxExcel
     global Unidad
     TxExcel+=1
     RutExcel=rut
     InmobiliariaExcel=inmobiliaria
     Unidad=unidad


     

def reportarError(mensaje,nomHoja,Rol,Unidad):
    # Cargar el archivo Excel
    fecha_actual=datetime.now()
    fecha_formateada = fecha_actual.strftime('%d/%m/%Y %H:%M:%S')
    archivo_excel = "Excel/Formato Reporte.xlsx"
    libro = openpyxl.load_workbook(archivo_excel)

    # Seleccionar la hoja en la que deseas escribir
    if nomHoja==True:
        hoja = libro.get_sheet_by_name("Errores")
    else:
        hoja = libro.get_sheet_by_name("Correctos")
    # Encontrar la última fila ocupada en una columna específica (por ejemplo, columna A)
    ultima_fila = hoja.max_row + 1

    # Lista de valores a insertar en las celdas
    valores_a_insertar = [str(TxExcel),str(fecha_formateada),str(RutExcel),str(InmobiliariaExcel),str(Rol),str(Unidad),str(mensaje)]
    # Escribir en las celdas vacías después de la última fila ocupada en la columna A
    for i, valor in enumerate(valores_a_insertar):
        hoja.cell(row=ultima_fila, column=i + 1, value=valor)

    # Guardar los cambios en el archivo
    libro.save(archivo_excel)
    libro.close()

def recorrerFilasDescargas(carpeta,driver,rol,porRut,nomInmobiliaria,unidad):
    row=0
    consecutivo=0
    tabledata=None   
    tabledata=txtscraping(carpeta)
    save=False
    filtroCuota=filtrarCuota()
    intFiltro=int(filtroCuota[0:1])
    intAnioFiltro=int(filtroCuota[2:])
    for celda in tabledata:
        row=row+1
        save=False
        try:
                CUOTA = celda.get('CUOTA')
                VALOR =  celda.get('VALOR')
                print(CUOTA)
                try:
                    strCuota=str(CUOTA)
                    cutCuota=strCuota[0:1]
                    intCuota=int(cutCuota)
                    anioCuota=int(strCuota[2:])
                except:
                     intCuota=99
                



                if intCuota<=intFiltro or anioCuota<intAnioFiltro:
                    #Descarga el PDF
                    if porRut==True:
                        consecutivo=consecutivo+1
                        print("El consecutivo es " + str(consecutivo))
                        descarga=driver.find_element(By.XPATH,"//table[@id='example']//tr["+str(row)+"]//td[3]")
                        descarga.click()
                        time.sleep(3)
                        while save==False:
                            #Ciclo para asegurar de que guarde el PDF en la carpeta con exito
                            try:
                                savepdf(nomInmobiliaria,str(consecutivo),CUOTA,str(rol))
                                save=True
                            except:
                                save=False
                        reportarError("Cuota "+str(CUOTA)+" rescatada con exito",False,rol,unidad)
                    else:
                        consecutivo=consecutivo+1
                        print("El consecutivo es " + str(consecutivo))
                        #clickweb("//TABLE[@id='example']//tr["+str(row)+"]//td[3]")
                        descarga=driver.find_element(By.XPATH,"//table[@id='example']//tr["+str(row)+"]//td[3]")
                        descarga.click()
                        savepdfRol(nomInmobiliaria,str(consecutivo),CUOTA,str(rol))
                        reportarError("Cuota "+str(CUOTA)+" rescatada con exito",False,rol,unidad)
                else:
                    #No es una cuota valida
                    print("La cuota ["+CUOTA+"] no se descarga")
                              
        except Exception as e:
             print("Falló", e)
             pass
        finally:
             pass


def savepdfRol(carpeta,consecutivo,cuota,rol):
 base=Pyasset(asset="base")
 txt=base+carpeta
 switch=False
 salida="Cupon de pago "+str(consecutivo)
 if str(consecutivo)=="1":
     consecutivo="1"

 
 try:
        file = open(txt+"\\"+salida)
        print(file) # File handler
        file.close()
 except:

    
    origen=txt+"\\"+salida+".pdf"
    destino=txt+"\\"+"Cupon de pago "+str(rol)+" "+str(cuota)+".pdf"
    intentos=0
    while switch==False:
        if os.path.exists(origen):
             try:os.remove(origen)
             except:pass
        if os.path.exists(destino):
            print("El Archivo ya existe")
            switch=True
        else:
            library.click("name:imprimirAr")
            time.sleep(4.5)
            library.send_keys(keys="{CTRL}P")
            time.sleep(1)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{Enter}")
            time.sleep(0.5)
            library.send_keys(keys="g")
            time.sleep(0.5)
            library.send_keys(keys="{Enter}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{TAB}")
            time.sleep(0.5)
            library.send_keys(keys="{Enter}")
            time.sleep(0.5)


            library.send_keys(keys=txt)
            time.sleep(5)
            library.send_keys(keys="{Enter}")
            time.sleep(2)
            #library.send_keys(keys="{Alt}N")
            time.sleep(2)
            library.send_keys(keys="{CTRL}A")
            time.sleep(2)
            library.send_keys(keys=str(salida))
            time.sleep(3)
            library.send_keys(keys="{Enter}")
            time.sleep(2)
            switch=cambionombre(origen, destino,str(rol),str(intentos), cuota)
        intentos+=1
    library.send_keys(keys="{Enter}")
    time.sleep(1)
    library.send_keys(keys="{Esc}")
    time.sleep(1)
    library.click("name:imprimirAr")
    time.sleep(1)
    library.send_keys(keys="{Ctrl}W")
    


def recorriendoFormatoSolicitud(carpeta,hoja):
    row=0
    
    tabledata=txtscraping(carpeta)
    try:
        for celda in tabledata:
            row=row+1  
            consecutivo=str(row)  
            CUOTA = celda.get('CUOTA')            
            VALOR=  celda.get('VALOR')
            si=str(CUOTA).find("-")

            if si == -1:                
                print("-----------------------------------------------------------------------")
            else:
                print("consultado hoja : "+hoja)
                print("consultado Cuota : "+str(CUOTA))
                print("consultado Monto : "+str(VALOR))
                row=int(row-1  )
                
               
                 
                row=row+1      
    except:
        pass
    finally:
            row=0
        
            tabledata=txtscraping(carpeta)
   
            for celda in tabledata:
                row=row+1  
                consecutivo=str(row)  
                CUOTA = celda.get('CUOTA')            
                VALOR=  celda.get('VALOR')
                si=str(CUOTA).find("-")

                if si == -1:                
                    print("-----------------------------------------------------------------------")
                else:
   
                    print("consultado hoja : "+hoja)
                    print("consultado Cuota : "+str(CUOTA))
                    print("consultado Monto : "+str(VALOR))
                    row=int(row-1  )
                
                    
                    row=row+1      
        
def validacion():
    validacion= browser.get_text("//DIV[@class='dentro_letra'][text()='Contribuciones']")
    if validacion == 'Contribuciones': print("ingresando a "+validacion) 
    return validacion



def agregarRolNoEncontrado(rol1, rol2):
    with open('Roles No encontrados.txt', 'a') as archivo:
            # Escribir el número en el archivo, seguido de un salto de línea
            archivo.write(str(rol1)+' '+str(rol2)+'\n')
    # Informar que se han escrito los números en el archivo
    print("Se agrego el Rol No encontrado")

def abrirNavegadorRut(userRut,userPass):
     #Ingresa a la página de contribuciones con RUT
     esperar=True
     c=0
     driver = webdriver.Firefox()
     driver.set_page_load_timeout(30)
     try:
        driver.get("https://www.tgr.cl/contribuciones/")  
     except:
        driver.quit()
     try:
         driver.maximize_window()
         elemento=driver.find_element(By.ID, "tgr-wp-contribuciones-pagopresencial-pagoasociadosrut")
         elemento.click()
         time.sleep(10)
         elemento=driver.find_element(By.ID, "id-button-idp-claveTributaria")
         elemento.click()
         time.sleep(5)
         while esperar==True:
            if c==10:
                break
            try:
                elemento=driver.find_element(By.ID, "user_rut")
                esperar=False
                time.sleep(1)
            except:
             c+=1
             esperar=True
             time.sleep(6)
         elemento=driver.find_element(By.ID, "user_rut")
         elemento.send_keys(userRut)#Cambiar por credenciales
         time.sleep(1)
         elemento=driver.find_element(By.ID, "user_pass")
         elemento.send_keys(userPass)#Cambiar por credenciales
         time.sleep(1)
         elemento=driver.find_element(By.ID,"bt_ingresar")
         elemento.click()
         time.sleep(25)
     
         #Busca el iframe que contiene el formulario
         esperar=True
         c=0
         while esperar==True:
            if c==10:
                break
            try:
                div=driver.find_element(By.ID,"tgr-sp-contenedor-iframe")
                iframes=div.find_element(By.TAG_NAME,"iframe")
                esperar=False
                time.sleep(1)
            except:
             c+=1
             esperar=True
             time.sleep(6)

         #Cambia al iframe que contiene el formulario
         driver.switch_to.frame(iframes)
     except:
         driver.switch_to.default_content()
         driver.quit()
         return driver
         
     
     return driver


def busquedaRol(driver,ruta,rol1,rol2,rolCompleto,nomInmobiliaria,unidad):
     try:
         #Busca el formulario
         elementoForm=driver.find_element(By.ID,"formc")

         #Escribe el rol a buscar 
         elementoDiv=elementoForm.find_element(By.ID,"example_filter")
         elemento=elementoDiv.find_element(By.TAG_NAME,"input")
         elemento.send_keys(rolCompleto)
         time.sleep(3)
     except:
         driver.switch_to.default_content()
         driver.quit()
         
     try:    
         tabla=elementoForm.find_element(By.ID,"example")
         tablaVacia=tabla.find_element(By.XPATH,'//td[@class="dataTables_empty"]')
         print("No se han encontrado Registros del rol: "+ rolCompleto)
         agregarRolNoEncontrado(rol1,rol2)
         elemento.clear()
     except:
         try:
             #Selecciona generar cupon de pago
             elemento=elementoForm.find_element(By.XPATH,'//input[@value="PortalContribPresencial"]')
             elemento.click()
             time.sleep(3)

             #Marca el check con el rol buscado
             elementoTable=elementoForm.find_element(By.ID,"example")
             elemento=elementoTable.find_element(By.ID,'selecctall')
             elemento.click()
             time.sleep(1)

             #Da click en aceptar
             elemento=driver.find_element(By.NAME,"enviar2")
             driver.execute_script("arguments[0].scrollIntoView();", elemento)
             elemento.click()
             time.sleep(5)
             elementoForm=driver.find_element(By.ID,"formc")
             driver.execute_script("arguments[0].scrollIntoView();", elementoForm)
         except:
              driver.quit()

         try: # Validando si la tabla funciona
             valida=elementoForm.find_element(By.XPATH,"//td[@class='celdaContenido2  sorting_1'][text()='NO POSEE DEUDAS DE CONTRIBUCIONES PARA LA CONSULTA REALIZADA']")
             textovalidacion='NO POSEE DEUDAS DE CONTRIBUCIONES PARA LA CONSULTA REALIZADA'
             reportarError("No Presenta Cuotas",False,str(rol1)+"-"+str(rol2),str(unidad))  
             if valida == textovalidacion:
                 print(textovalidacion)
                 #Busca el boton para volver
                 time.sleep(2)
                 elemento=elementoForm.find_element(By.NAME,"Ingresar")
                 driver.execute_script("arguments[0].scrollIntoView();", elemento)
                 elemento.click()
                 time.sleep(5)
                 #Escribe el rol a buscar
                 elementoForm=driver.find_element(By.ID,"formc")
                 elementoDiv=elementoForm.find_element(By.ID,"example_filter")
                 elemento=elementoDiv.find_element(By.TAG_NAME,"input")
                 elemento.send_keys(rolCompleto)
                 time.sleep(3)
                 #Desmarca el check con el rol buscado
                 elementoTable=elementoForm.find_element(By.ID,"example")
                 elemento=elementoTable.find_element(By.ID,'selecctall')
                 elemento.click()
                 time.sleep(1)
                 elemento=elementoDiv.find_element(By.TAG_NAME,"input")
                 elemento.clear()
         except:
             #proceso de consulta
             generarLogScraping(driver,ruta)
             pdfrol=str(rol1)+"-"+str(rol2)
             recorrerFilasDescargas(ruta,driver,str(pdfrol),True,nomInmobiliaria,unidad)
             #Busca el boton para volver
             time.sleep(2)
             elemento=elementoForm.find_element(By.NAME,"Ingresar")
             driver.execute_script("arguments[0].scrollIntoView();", elemento)
             elemento.click()
             time.sleep(5)
             #Escribe el rol a buscar
             elementoForm=driver.find_element(By.ID,"formc")
             elementoDiv=elementoForm.find_element(By.ID,"example_filter")
             elemento=elementoDiv.find_element(By.TAG_NAME,"input")
             elemento.send_keys(rolCompleto)
             time.sleep(3)
             #Desmarca el check con el rol buscado
             elementoTable=elementoForm.find_element(By.ID,"example")
             elemento=elementoTable.find_element(By.ID,'selecctall')
             elemento.click()
             time.sleep(1)
             elemento=elementoDiv.find_element(By.TAG_NAME,"input")
             elemento.clear()
             




def generarLogScraping(driver,carpeta):
    # Encuentra la tabla por su selector CSS (ajusta el selector según tu caso)
    elementoForm=driver.find_element(By.ID,"formc")
    tabla = elementoForm.find_element(By.XPATH,"//table[@id='example']")

    # Encuentra todas las filas de la tabla
    filas = tabla.find_elements(By.TAG_NAME,'tr')

    # Abre un archivo de texto en modo escritura
    with open("Log Scraping/"+carpeta+".txt", 'w', encoding='utf-8') as archivo:

    # Itera a través de las filas y guarda el texto de las celdas en el archivo
        for fila in filas:
            celdas = fila.find_elements(By.TAG_NAME,'td')
            fila_datos = [celda.text for celda in celdas]
            fila_texto = "\t".join(fila_datos)  # Separar celdas con tabulaciones
            archivo.write(fila_texto + '\n')  # Escribe la fila en el archivo


def generarLogScrapingRol(driver,carpeta):
    # Encuentra la tabla por su selector CSS (ajusta el selector según tu caso)
    tabla = driver.find_element(By.XPATH,"//table[@id='example']")

    # Encuentra todas las filas de la tabla
    filas = tabla.find_elements(By.TAG_NAME,'tr')

    # Abre un archivo de texto en modo escritura
    with open("Log Scraping/"+carpeta+".txt", 'w', encoding='utf-8') as archivo:

    # Itera a través de las filas y guarda el texto de las celdas en el archivo
        for fila in filas:
            celdas = fila.find_elements(By.TAG_NAME,'td')
            fila_datos = [celda.text for celda in celdas]
            fila_texto = "\t".join(fila_datos)  # Separar celdas con tabulaciones
            archivo.write(fila_texto + '\n')  # Escribe la fila en el archivo


def probando():
    current_working_directory = os.getcwd()
    capsolver_extension_path = current_working_directory + "\\capsolver\\"
    chrome_service = Service()
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(
            "--load-extension={0}".format(capsolver_extension_path))
    # Desactivar la advertencia de que los archivos pueden ser peligrosos



    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
    driver.set_page_load_timeout(60)
    try:
        driver.get("https://www.africau.edu/images/default/sample.pdf")  
    except:
        driver.quit()
    library.send_keys(keys="{CTRL}S")    
    time.sleep(2)
    library.send_keys(keys="{Enter}")    
    time.sleep(59)

def navegacion(region,comuna,rol1,rol2,ruta,nomInmobiliaria,unidad):
    current_working_directory = os.getcwd()
    capsolver_extension_path = current_working_directory + "\\capsolver\\"
    chrome_service = Service()
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(
            "--load-extension={0}".format(capsolver_extension_path))


    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
    driver.set_page_load_timeout(60)
    try:
        driver.get("https://www.tesoreria.cl/ContribucionesPorRolWEB/muestraBusqueda?tipoPago=PortalContribPresencial")  
    except:
        driver.quit()

    driver.maximize_window()
    r=region.upper()
    elemento=driver.find_element(By.XPATH,"//select[@id='region']")
    elemento.click()
    time.sleep(3)
    elemento=driver.find_element(By.XPATH,"//option[text()='"+r+"']")
    elemento.click()
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//select[@id='comunas']")
    elemento.click()
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//option[text()='"+comuna+"']")
    elemento.click()
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//input[@id='rol']")
    elemento.send_keys(str(rol1))
    time.sleep(1)
    elemento=driver.find_element(By.XPATH,"//input[@id='subRol']")
    elemento.send_keys(str(rol2))
    time.sleep(3)
    xpath = '//input[@id="btnRecaptchaV3Envio" and @class="boton g-recaptcha"]'
    elemento=driver.find_element(By.ID,"btnRecaptchaV3Envio")
    driver.execute_script("arguments[0].scrollIntoView();", elemento)
    elemento.click()
    time.sleep(5)
    try:
        capsolver = (WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/div"
                                                                                              "/div[4]/div/div[2]"))))
        capsolver.click()
        time.sleep(10)
    except:
        print("No funciono el captcha")
        

    espera=True
    cantidad=0
    #tiempoespera()
    while espera==True:
        cantidad+=1
        if cantidad==7:
            espera=False
            raise ValueError("Error")
            break
        try:
            valida=driver.find_element(By.XPATH,"//div[@id='example_filter']")
            time.sleep(1)
            print("Ya cargó")
            espera=False
        except:
            espera=True
            print("-----------Cargando-------------")
            time.sleep(5)


    try: # Validando si la tabla funciona
        valida=driver.find_element(By.XPATH,"//td[@class='celdaContenido2  sorting_1'][text()='No se encontraron Deudas']")
        #valida=obtenerTexto("//TD[@class='celdaContenido2  sorting_1'][text()='No se encontraron Deudas']/self::TD")
        print("No hay registros de este Rol")
        reportarError("No Presenta Cuotas",False,str(rol1)+"-"+str(rol2),str(unidad))  
    except:
        try:# proceso de consulta
                    #tabla = None
                    #tabla =extraertablita()
                    #export(ruta,tabla)
                    generarLogScrapingRol(driver,ruta)
                    pdfrol=str(rol1)+"-"+str(rol2)
                    recorrerFilasDescargas(ruta,driver,str(pdfrol),False,nomInmobiliaria,str(unidad))
                    #recorrerFilasDescargas(ruta,tabla,str(pdfrol),False,nomInmobiliaria)
        except:# proceso de consulta reintento #1
            driver.quit()
            #cerraNavegador()
            raise ValueError("Error")

    finally:
         driver.quit()
         return driver
         #cerraNavegador() 
                           

def savepdf(carpeta,consecutivo,cuota,rol):
 base=Pyasset(asset="base")
 txt=base+carpeta
 switch=False
 alter=False
 salida="Cupon de pago "+str(consecutivo)
 if str(consecutivo)=="1":
     consecutivo="1"

 
 try:
        file = open(txt+"\\"+salida)
        print(file) # File handler
        file.close()
 except:

    library.click("name:imprimirAr")
    time.sleep(4.5)
    library.send_keys(keys="{CTRL}S")   
    time.sleep(4)
    
    origen=txt+"\\"+salida+".pdf"
    destino=txt+"\\"+"Cupon de pago "+str(rol)+" "+str(cuota)+".pdf"
    intentos=0
    while switch==False:
        if os.path.exists(origen):
             try:os.remove(origen)
             except:pass
        if os.path.exists(destino):
            print("El Archivo ya existe")
            switch=True
        else:
            library.click("name:imprimirAr")
            time.sleep(4.5)    
            library.send_keys(keys="{CTRL}S")    
            time.sleep(2)

            if alter==True:
                consecutivo="2"
            if str(consecutivo)==str("1"):
                library.send_keys(keys=txt)
                time.sleep(5)
                library.send_keys(keys="{Enter}")
                time.sleep(2)
                library.send_keys(keys="{CTRL}A")
                time.sleep(2)
                library.send_keys(keys=str(salida))
                time.sleep(3)
                library.send_keys(keys="{Enter}")
                switch=cambionombre(origen, destino,str(rol),str(intentos))
                if switch==False:
                     alter=True

            if str(consecutivo)!=str("1"):
                if alter==True:
                     library.send_keys(keys="{CTRL}S")    
                     time.sleep(2)
                consecutivo="1"
                alter=False
                library.send_keys(keys=str(salida))
                time.sleep(3)
                library.send_keys(keys="{Enter}")
      
                switch=cambionombre(origen, destino, str(rol),str(intentos))
        intentos+=1
    library.send_keys(keys="{Enter}")
    time.sleep(1)
    library.send_keys(keys="{Esc}")
    time.sleep(1)
    library.click("name:imprimirAr")
    time.sleep(1)
    library.send_keys(keys="{Ctrl}W")
        
   
def txtscraping(carpeta):
    listSCRAPIADO = []  # Inicializamos la lista que contendrá los datos

    with open('Log Scraping/' + carpeta + ".txt", "r") as f:
        for x in f:
            x = x.strip()  # Eliminamos espacios en blanco al principio y al final

            if not x:  # Saltar líneas en blanco
                continue

            # Dividir la línea en palabras separadas por espacios
            palabras = x.split()

            # Verificamos que haya suficientes palabras en la línea
            if len(palabras) >= 5:
                CUOTA = palabras[0]
                VALOR = palabras[1].replace(",", "")
                NRO_FOLIO = palabras[2].replace(",", "")
                VENCIMIENTO = palabras[3].replace(",", "")
                TOTA_PAGAR = palabras[4].replace(",", "")

                # Agregar los datos a la lista
                listSCRAPIADO.append({
                    'CUOTA': CUOTA,
                    'NRO FOLIO': NRO_FOLIO,
                    'VALOR': VALOR,
                    'VENCIMIENTO': VENCIMIENTO,
                    'TOTA A PAGAR': TOTA_PAGAR
                })

    return listSCRAPIADO
    

       
    
def export(Carpeta,tabla):
     
     datosscrap=str(tabla) 
     outmensaje=datosscrap
     outmensaje=outmensaje.replace("VALOR"," ")
     outmensaje=outmensaje.replace("CUOTA"," ")
     outmensaje=outmensaje.replace("VALOR CUOTA"," " )
     outmensaje=outmensaje.replace("NRO FOLIO"," " )
     outmensaje=outmensaje.replace("VENCIMIENTO"," " )
     outmensaje=outmensaje.replace("TOTAL A PAGAR"," " )
     outmensaje=outmensaje.replace("EMAIL"," " )
     outmensaje=outmensaje.replace("DESCARGAR"," " )
     outmensaje=outmensaje.replace("""CUOTA
                                    VALOR CUOTA
                                    NRO FOLIO
                                    VENCIMIENTO
                                    TOTAL A PAGAR
                                    EMAIL
                                    DESCARGAR"""," " )

     try:
        file = open("Log Scraping/"+Carpeta+".txt")
        print(file) # File handler
        file.close()
       
     except:
        print("Archivo no existe se genera uno nuevo  "+ "Log Scraping/"+Carpeta+".txt")
        nom="Log Scraping/"+Carpeta+".txt"     
        f = open(nom, "a")
        f.write(outmensaje)
        f.close() 
                      
def cambionombre(origen, destino, rol,nIntentos,cuota):

        archivo = origen
        global Unidad
        nombre_nuevo = destino
        print("archivo → "+ archivo )
        print("Destino → "+ nombre_nuevo)
        try:
            os.rename(archivo, nombre_nuevo)
            if os.path.exists(nombre_nuevo):
                 print("PDF guardado con exito")
                 return True
            else:
                print("Ocurrio un error al guardar el PDF, reintentando (Else)")
                library.send_keys(keys="{Enter}")
                library.send_keys(keys="{Esc}")
                reportarError("Fallo al guardar la cuota "+cuota+" - Reintento N°"+str(nIntentos),True,rol,Unidad)
                return False
        except:
             print("Ocurrio un error al guardar el PDF, reintentando (Except)")
             time.sleep(1)
             library.send_keys(keys="{Enter}")
             time.sleep(1)
             library.send_keys(keys="{Esc}")
             reportarError("Fallo al guardar la cuota "+cuota+" - Reintento N°"+str(nIntentos),True,rol,Unidad)
             return False




def Resumen():
    lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
    lib.read_worksheet("Resumen")                                              #nombre de la hoja
    dtresumen=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data
    return dtresumen

def master():
   lib.open_workbook("Data\Master.xlsx")      #ubicacion del libro
   lib.read_worksheet("Listado")       #nombre de la hoja
   DtMaster=lib.read_worksheet_as_table(name='Listado',header=True, start=1).data

   return DtMaster

def diligenciarResumen(h,carpeta):
    dtcon=txtscraping(carpeta)
   
       #ahora = datetime.now()
       #consulta=str(ahora.year)
    consulta="2023"
    
     
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR
        
                lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
                lib.read_worksheet("Resumen")                                              #nombre de la hoja
                libroresumen=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data    

                cantidad=lib.find_empty_row()

                #Ingresamos los valores 
                for celda in range(cantidad):
                
                    Numero=lib.get_cell_value(2+celda,"A")
                    if Numero==h:
                            lib.set_cell_value(2+celda,"E",str(v))
                            lib.set_cell_value(2+celda,"f","pago contribucciones "+str(cu))
                            lib.save_workbook() 


                                
def formatosolicitusd(h,totalR,numeroSolicitud):

    total=float(totalR)

    fecha_actual = datetime.now()

    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')

    #ahora = datetime.now()
    #consulta=str(ahora.year)
    #consulta=fecha_actual.year
    switch=False
    consulta=filtrarCuota()
    origen='Data\\Formato Solicitud Pago.xlsx'

    datac=Resumen()

    for x in datac:
        if x[0]==h:
                        
            lib.open_workbook(origen)      
            lib.read_worksheet("Solicitud")                                              
            libroresumen=lib.read_worksheet_as_table(name='Solicitud',header=True, start=1).data
                        
            lib.set_cell_value(8,"D",str(x[8]))

            lib.set_cell_value(6,"H",str(fecha_formateada))
            lib.set_cell_value(10,"D","Enrique Carrasco")
            lib.set_cell_value(12,"D",str(x[3]))
            lib.set_cell_value(12,"H",str(x[2]))
            lib.set_cell_value(14,"C",int(total), fmt="[$$-es-CL]#,##0")
            lib.set_cell_value(20,"C","Teatinos 28, Santiago")
            lib.set_cell_value(22,"D","Pago contribucciones "+str(consulta))
            lib.set_cell_value(24,"D","Pago contribucciones "+str(consulta))
            lib.set_cell_value(26,"D",str(x[12]))
            lib.set_cell_value(28,"D",str(x[12]))
            lib.set_cell_value(30,"D",str("Contribucciones"))
            try:
                 os.mkdir("Formato Solicitud\\Solicitudes De Pago "+str(x[3]))
            except:pass
            while switch==False:
                 try:     
                    destino="Formato Solicitud\\Solicitudes De Pago "+str(x[3])+"\\Solicitud Pago "+str(x[3])+" "+str(numeroSolicitud)+".xlsx"
                    switch=True
                    lib.save_workbook(destino)
                 except:
                    switch=False
            lib.close_workbook()
                        
def diligenciarhojas(h,carpeta,REGION,COMUNA,ROLMATRIZ,RUT,INMOBILIARIA,rol1,rol2):
    dtcon=txtscraping(carpeta)
    R=0
    celda=0

    for txt in dtcon:
         celda=1+celda

    print("el total de celdas es → "+str(celda))
    
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
            print(CUOTA)           
            VALOR=  txt.get('VALOR')
        
            lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
            lib.read_worksheet(str(h))                                                  #nombre de la hoja
            libroresumen=lib.read_worksheet_as_table(name=str(h),header=True, start=1).data    
            
            R=1+R 
                       
            lib.set_cell_value(6+R,"B",RUT) 
            lib.set_cell_value(6+R,"C",INMOBILIARIA)
            lib.set_cell_value(6+R,"D",REGION)
            lib.set_cell_value(6+R,"E",COMUNA)
            lib.set_cell_value(6,"H","Monto")
            lib.set_cell_value(5+R,"H",VALOR,fmt="0.00")
            lib.set_cell_value(6+R,"D",REGION)
            lib.set_cell_value(6+R,"E",COMUNA)
            lib.set_cell_value(6+R,"F",ROLMATRIZ)                   
            lib.save_workbook()
            
    print("el total de R es → "+str(R))
    R=0       
    lib.clear_cell_range("B16:H77")        
    for txt in dtcon:
            CUOTA = txt.get('CUOTA')                        
            VALOR=  txt.get('VALOR')
            R=1+R
            VO=lib.get_cell_value(5+R,"H")
            if VO is None:
                print(VO)
                lib.set_cell_value(5+R,"G"," ")
                lib.set_cell_value(5+R,"F"," ")
                lib.set_cell_value(5+R,"E"," ")
                lib.set_cell_value(5+R,"D"," ")
                lib.set_cell_value(5+R,"C"," ")
                lib.set_cell_value(5+R,"B"," ")
                break
            else:
                lib.set_cell_value(5+R,"G",CUOTA,fmt="0")

        
       
            
    #lib.set_cell_value(7+(R+2),"G","Total") 
    #lib.set_cell_formula("H17","=SUMA(H7:H16)",True)
   
    lib.set_cell_value(6,"H","Monto") 
    lib.save_workbook("Salida\\Resumen_Contribuciones_Terreno_2023.xlsx")#"Salida\\Resumen_Contribuciones_Terreno_2023.xlsx"
    lib.close_workbook ()      

def bakup():
     
     print("Realizamos el bakup")
     origen='Data\\BACKUP\\Resumen_Contribuciones_Terreno_2023.xlsx'         
     destino="Data\\Resumen_Contribuciones_Terreno_2023.xlsx"
     shutil.copy(origen,destino )
     origen='Excel\\Formato Reporte\\Formato Reporte.xlsx'         
     destino="Excel\\Formato Reporte.xlsx"
     shutil.copy(origen,destino )

def creacioncarpetas (carpeta):
    os.mkdir('PDF/'+carpeta)    
    print("creacion de carpetas  PDF/"+carpeta) 

def Macros (h):
    lib.open_workbook("Data\Macro TGR.xlsm")      
    lib.read_worksheet("MACRO")                                                                     
    libroresumen=lib.read_worksheet_as_table(name="MACRO",header=True, start=1).data 

    lib.set_cell_value(3,"B",str(h))
    lib.save_workbook()
    lib.close_workbook()
    time.sleep(10)

    app.open_application(visible=True)
    try:
         library.click("name:Cerrar")
    except:
         pass

    app.open_workbook('Data\Macro TGR.xlsm')
    app.set_active_worksheet(sheetname="MACRO")
    time.sleep(5)
    app.run_macro("Main")
    time.sleep(5)
    app.save_excel()
    app.quit_application()

def totalMacro(h):
    #BORRAR?
    lib.open_workbook("Data\Resumen_Contribuciones_Terreno_2023.xlsx")      
    lib.read_worksheet(h)                                                                     
    libroresumen=lib.read_worksheet_as_table(name=str(h),header=True, start=1).data 

    TOTAL =lib.get_cell_value(20,"H")

    lib.save_workbook()
    lib.close_workbook()
    return TOTAL

def formatoTotal(h,carpeta):


    totalv=int(totalMacro(h))
    print(str(totalv))
    
    dtcon=txtscraping(carpeta)

    fecha_actual = datetime.now()

    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')
    

    #ahora = datetime.now()
    #consulta=str(ahora.year)
    consulta="2018"
    
      
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR 
       
                destino="Formato Solicitud\\"+carpeta +" " +" Cuota " + str(cu) + " Formato Solicitud Pago.xlsx"
                #shutil.copy(origen,destino )

                datac=Resumen()

                for x in datac:
                    if x[0]==h:

                        lib.open_workbook(destino)      
                        lib.read_worksheet("Solicitud")                                                                     
                        libroresumen=lib.read_worksheet_as_table(name='Solicitud',header=True, start=1).data 

                    
                        lib.set_cell_value(14,"C",int(totalv), fmt="0.00")
                        

            

    lib.save_workbook()
    lib.close_workbook()

def fGuardar(h,carpeta):

    dtcon=txtscraping(carpeta)

    fecha_actual = datetime.now()

    fecha_formateada = fecha_actual.strftime('%d/%m/%Y')

    #ahora = datetime.now()
    #consulta=str(ahora.year)
    consulta="2018"
    
      
    for txt in dtcon:
            CUOTA = txt.get('CUOTA') 
                       
            VALOR=  txt.get('VALOR')
            if str(CUOTA)[2:]==consulta : 
                cu = CUOTA             
                v = VALOR 
     
    destino="Formato Solicitud\\"+carpeta +" " +" Cuota " + str(cu) + ".xlsm"

    lib.open_workbook("Data\Resumen_Contribuciones_Terreno_2023.xlsm")      
    lib.read_worksheet("Solicitud")                                                                     
    libroresumen=lib.read_worksheet_as_table(name='Solicitud',header=True, start=1).data 

    lib.set_cell_value(1,"k",int(h))

    lib.save_workbook(destino)
    lib.close_workbook()

def ResumenFinal ():
     lib.open_workbook('Data\Resumen_Contribuciones_Terreno_2023.xlsx')        #ubicacion del libro
     lib.read_worksheet('Resumen')       #nombre de la hoja
     lista=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data

     ultimaFila= lib.find_empty_row()

     for celda in range(ultimaFila):
         TOTAL= lib.get_cell_value(2+int(celda),"E")
         if TOTAL == "=+'1'!$H$9":
              print("True "+str(TOTAL))
         else:
              print("false "+str(TOTAL))
              HOJA=lib.get_cell_value(2+int(celda),"A")
              
              lib.read_worksheet(str(HOJA))
              tablaTotal=lib.get_cell_value(20,"H")
             

              lib.read_worksheet('Resumen')
              lib.set_cell_value(2+int(celda),"E",int(tablaTotal))



              lib.save_workbook()
              lib.close_workbook()

def limpiarResumen():
     
     
     lib.open_workbook("Data\\Resumen_Contribuciones_Terreno_2023.xlsx")      #ubicacion del libro
     lib.read_worksheet("94")                                                 #nombre de la hoja
     libroresumen=lib.read_worksheet_as_table(name="94",header=True, start=1).data  
      

     lib.clear_cell_range("G7:G1000")
     rango="B{}:H{}"
     
     #Comparaciones 
     item1=lib.get_cell_value(7,"H")
     item2=lib.get_cell_value(8,"H")
     item3=lib.get_cell_value(9,"H")
     item4=lib.get_cell_value(10,"H")
     item5=lib.get_cell_value(11,"H")
     item6=lib.get_cell_value(12,"H")
     item7=lib.get_cell_value(13,"H")
     item8=lib.get_cell_value(14,"H")
     item9=lib.get_cell_value(16,"H")
     item10=lib.get_cell_value(17,"H")
      
     busquedad=0
     for x in range(1000):           
            Cels=str(x+8)
                   
            if item1==lib.get_cell_value(7,"H"):
               busquedad=1+busquedad   
            elif busquedad>1:      
               lib.clear_cell_range(rango.format(Cels,Cels))

     busquedad=0
     for x in range(1000):           
            Cels=str(x+9)
                   
            if item1==lib.get_cell_value(8,"H"):
               busquedad=1+busquedad   
            elif busquedad>1:      
               lib.clear_cell_range(rango.format(Cels,Cels))




     lib.save_workbook()
     lib.close_workbook()

def salida():
     print("Realizamos la salida ")
     origen='Data\Resumen_Contribuciones_Terreno_2023.xlsx'         
     destino="Salida\Resumen_Contribuciones_Unidades_2023.xlsx"
     shutil.copy(origen,destino )
     origen="Excel\Formato Reporte.xlsx"
     destino="Salida\Formato Reporte.xlsx"
     shutil.copy(origen,destino )


