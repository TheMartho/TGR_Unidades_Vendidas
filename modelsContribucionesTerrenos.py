from RPA.Tables import Tables
from RPA.Excel.Files import Files
from RPA.Excel.Application import Application
lib = Files()
import modelsUnidadesvendidas
import modelsSolicitudPago
library = Tables()
import defRPAselenium
import pandas as pd
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import openpyxl.styles


def agregarBordes(data, numeroInmobiliaria,uFila):
    excel_file = load_workbook('Data\Resumen_Contribuciones_Terreno_2023.xlsx')
    hoja = excel_file([str(numeroInmobiliaria)])
    rango= hoja['B7:H'+str(uFila)]

    for row in rango:
        for cell in row:
            cell.font = openpyxl.styles.Font()
            cell.border = openpyxl.styles.Border()
            cell.fill = openpyxl.styles.PatternFill()
            cell.alignment = openpyxl.styles.Alignment()
            cell.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    excel_file.save('Data\Resumen_Contribuciones_Terreno_2023.xlsx')




"""def masterResumen():
    lib.open_workbook('Data\Resumen_Contribuciones_Terreno_2023.xlsx')       
    lib.read_worksheet("Resumen")       
    DtableFinal=lib.read_worksheet_as_table(name="Resumen",header=True, start=1).data
    return DtableFinal"""

def resumenHoja(inmobiliaria):
    strInmobiliaria=str(inmobiliaria)
    lib.open_workbook('Data\Resumen_Contribuciones_Terreno_2023.xlsx')
    lib.read_worksheet(strInmobiliaria)
    DtableFinal=lib.read_worksheet_as_table(name=strInmobiliaria,header=True, start=1).data
    return DtableFinal

def reflejarTotal(inmobiliaria,numeroInmobiliaria):
    lib.open_workbook('Data\Resumen_Contribuciones_Terreno_2023.xlsx')
    lib.read_worksheet('Resumen')
    DtableFinal=lib.read_worksheet_as_table(name='Resumen',header=True, start=1).data
    consulta=defRPAselenium.filtrarCuota()
    valor="=+'"+str(numeroInmobiliaria)+"'!$L$8"
    fila=1
    for x in DtableFinal:
        inmobiliariaActual=str(lib.get_cell_value(fila,"D"))
        if inmobiliaria == inmobiliariaActual:
            lib.set_cell_value(fila,"E",valor)
            lib.set_cell_value(fila,"F","Pago contribucciones "+str(consulta))
            lib.save_workbook()
            lib.close_workbook()
            break
        fila+=1




def insertarResumen(consolidado,totalDefinitivo,inmobiliaria):
    numeroInmobiliaria=modelsSolicitudPago.getH(consolidado)
    resumenHoja(numeroInmobiliaria)
    u=modelsSolicitudPago.excelUnidadesVendidas(inmobiliaria[0:31])
    if u is None:
        print("No presenta cuotas")
    else:
        lib.set_cell_value(6,"B","INMOBILIARIA")
        lib.set_cell_value(6,"C","REGION")
        lib.set_cell_value(6,"D","COMUNA")
        lib.set_cell_value(6,"E","ROL MATRIZ")
        lib.set_cell_value(6,"F","ROL UNIDAD")
        lib.set_cell_value(6,"G","TIPO PRODUCTO")
        lib.set_cell_value(6,"H","NUMERO")
        lib.set_cell_value(6,"I","CUOTA")
        lib.set_cell_value(6,"J","TOTAL A PAGAR")


        contadorFila=6
        for fila in u:
            contadorFila+=1

            
            nombreConsolidado=fila['INMOBILIARIA']
            region=fila['REGION']
            comuna=fila['COMUNA']
            rolMatriz=fila['ROL MATRIZ']
            rolUnidad=fila['ROL UNIDAD']
            tipoProducto=fila['TIPO PRODUCTO']
            numero=fila['NUMERO']
            cuota=str(fila['CUOTA'])
            totalPagar=fila['TOTAL A PAGAR']
            if str(cuota)=="nan":
                cuota=""
            if str(totalPagar)=="nan" :
                totalPagar=""
            if str(nombreConsolidado)=="nan":
                nombreConsolidado=""
            if str(region)=="nan":
                region=""
            if str(comuna)=="nan":
                comuna=""
            if str(rolMatriz)=="nan":
                rolMatriz=""
            if str(rolUnidad)=="nan":
                rolUnidad=""
            if str(tipoProducto)=="nan":
                tipoProducto=""
            if str(numero)=="nan":
                numero=""



            lib.set_cell_value(contadorFila,"B",str(nombreConsolidado))
            lib.set_cell_value(contadorFila,"C",str(region))
            lib.set_cell_value(contadorFila,"D",str(comuna))
            lib.set_cell_value(contadorFila,"E",rolMatriz,fmt="0")
            lib.set_cell_value(contadorFila,"F",rolUnidad,fmt="0")
            lib.set_cell_value(contadorFila,"G",tipoProducto)
            lib.set_cell_value(contadorFila,"H",numero,fmt="0")
            lib.set_cell_value(contadorFila,"I",str(cuota))
            if '$' in str(totalPagar):
                totalPagar=str(totalPagar)
                lib.set_cell_value(contadorFila,"J",int(totalPagar[1:]),fmt="[$$-es-CL]#,##0")
            else:
                lib.set_cell_value(contadorFila,"J",totalPagar,fmt="[$$-es-CL]#,##0")

            #Ingresar TOTAL A PAGAR CON FORMATO PESO
            #lib.set_cell_value(contadorFila,"J",int(totalPagar),fmt="[$$-es-CL]#,##0")

            

            print("Insetando Fila N°"+str(contadorFila))

        lib.set_cell_value(8,"K",'Total')
        lib.set_cell_value(8,"L",totalDefinitivo,fmt="[$$-es-CL]#,##0")
        lib.save_workbook()
        lib.close_workbook()
        reflejarTotal(inmobiliaria,numeroInmobiliaria)
        






def task(inmobiliaria,totalDefinitivo,nombreConsolidado):
    print("Generando Resumen de "+ inmobiliaria)
    insertarResumen(nombreConsolidado,totalDefinitivo,inmobiliaria)

            